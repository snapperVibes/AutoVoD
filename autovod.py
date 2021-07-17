import datetime
from contextlib import redirect_stdout
from dataclasses import dataclass
from unittest.mock import MagicMock

import click
from moviepy.video.io.ffmpeg_tools import ffmpeg_extract_subclip
from openpyxl import load_workbook


@dataclass
class Match:
    p1: str
    p2: str
    _start: str
    _end: str

    def _to_time(self, time):
        if isinstance(time, datetime.time):
            raise TypeError(
                "Oops! It looks like the Excel spreadsheet views your time as an actual 'time' rather than text. "
            )
        h, m, s = map(int, time.split(":"))
        return (h * 3600) + (m * 60) + s

    @property
    def start_time(self) -> int:
        return self._to_time(self._start)

    @property
    def end_time(self) -> int:
        return self._to_time(self._end)

    @property
    def title(self) -> str:
        return f"{self.p1} vs {self.p2}"


@click.command("autovod")
@click.option("--data")
@click.option("--vod")
def cli(data, vod):
    # Load data
    wb = load_workbook(data)
    sheet = wb.active
    matches = []
    for row in sheet.iter_rows():
        matches.append(Match(*[data.value for data in row]))

    # Split the videos
    for match in matches:
        title = f"{match.title}.mp4"
        print(f"Exporting '{title}'")
        # By default, Moviepy outputs 3 lines of text for every clip.
        # This is a hack to get rid of them (so we can replace it with our own)
        with redirect_stdout(MagicMock()):
            ffmpeg_extract_subclip(
                vod, match.start_time, match.end_time, targetname=title
            )


if __name__ == "__main__":
    cli()
